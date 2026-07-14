"""
excel_export.py
Builds the region-grouped, subtotal + grand-total Excel report,
matching the layout of your original tracker PDF.
"""

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db import REGIONS

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GRANDTOTAL_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    "Region", "Site Code", "Plant", "DV Type", "# Total DV", "# DV Available for Today",
    "DV Utilised (Ord Recd)", "DV inroute to plant", "Unutilized DV - Balance",
    "Utilization % against Available DV",
]


def build_report_dataframe(latest_entries):
    """Returns a list of row dicts, grouped by region with subtotal + grand total rows.
    Each plant may appear more than once if it has entries under multiple DV Types --
    that's intentional segregation, not a duplicate."""
    rows = []
    grand = {"total_dv": 0, "dv_available": 0, "dv_utilised": 0, "dv_inroute": 0}

    for region in REGIONS:
        region_entries = [e for e in latest_entries if e["region"] == region]
        if not region_entries:
            continue

        sub = {"total_dv": 0, "dv_available": 0, "dv_utilised": 0, "dv_inroute": 0}
        for e in region_entries:
            balance = e["dv_available"] - e["dv_utilised"]
            util_pct = (e["dv_utilised"] / e["dv_available"]) if e["dv_available"] else 0
            opening_pct = (e["dv_available"] / e["total_dv"]) if e["total_dv"] else 0
            eff_util_pct = util_pct * opening_pct
            rows.append({
                "Region": e["region"], "Site Code": e["site_code"], "Plant": e["plant_name"],
                "DV Type": e.get("dv_type", ""),
                "Total DV": e["total_dv"], "DV Available": e["dv_available"],
                "DV Utilised": e["dv_utilised"], "DV Inroute": e["dv_inroute"],
                "Unutilized Balance": balance, "Utilization %": util_pct,
                "Opening DV %": opening_pct, "Effective Utilization %": eff_util_pct,
                "_row_type": "data",
            })
            for k in sub:
                sub[k] += e[k if k != "dv_available" else "dv_available"]

        sub_balance = sub["dv_available"] - sub["dv_utilised"]
        sub_util = (sub["dv_utilised"] / sub["dv_available"]) if sub["dv_available"] else 0
        sub_opening = (sub["dv_available"] / sub["total_dv"]) if sub["total_dv"] else 0
        sub_eff_util = sub_util * sub_opening
        rows.append({
            "Region": f"{region} Total", "Site Code": "", "Plant": "", "DV Type": "",
            "Total DV": sub["total_dv"], "DV Available": sub["dv_available"],
            "DV Utilised": sub["dv_utilised"], "DV Inroute": sub["dv_inroute"],
            "Unutilized Balance": sub_balance, "Utilization %": sub_util,
            "Opening DV %": sub_opening, "Effective Utilization %": sub_eff_util,
            "_row_type": "subtotal",
        })
        for k in grand:
            grand[k] += sub[k]

    if latest_entries:
        grand_balance = grand["dv_available"] - grand["dv_utilised"]
        grand_util = (grand["dv_utilised"] / grand["dv_available"]) if grand["dv_available"] else 0
        grand_opening = (grand["dv_available"] / grand["total_dv"]) if grand["total_dv"] else 0
        grand_eff_util = grand_util * grand_opening
        rows.append({
            "Region": "Pan India Total", "Site Code": "", "Plant": "", "DV Type": "",
            "Total DV": grand["total_dv"], "DV Available": grand["dv_available"],
            "DV Utilised": grand["dv_utilised"], "DV Inroute": grand["dv_inroute"],
            "Unutilized Balance": grand_balance, "Utilization %": grand_util,
            "Opening DV %": grand_opening, "Effective Utilization %": grand_eff_util,
            "_row_type": "grandtotal",
        })

    return rows


def _populate_worksheet(ws, entries, title):
    rows = build_report_dataframe(entries)

    ws.merge_cells("A1:L1")
    ws["A1"] = title
    ws["A1"].font = WHITE_BOLD
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Region", "Site Code", "Plant", "DV Type", "# Total DV", "# DV Available for Today",
               "DV Utilised (Ord Recd)", "DV inroute to plant", "Unutilized DV - Balance",
               "Utilization % against Available DV", "Opening DV %", "Effective Utilization %"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER

    r = 3
    for row in rows:
        values = [
            row["Region"], row["Site Code"], row["Plant"], row["DV Type"], row["Total DV"],
            row["DV Available"], row["DV Utilised"], row["DV Inroute"],
            row["Unutilized Balance"], row["Utilization %"],
            row["Opening DV %"], row["Effective Utilization %"],
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.border = BORDER
            if col_idx in (10, 11, 12):
                cell.number_format = "0%"
            if row["_row_type"] == "subtotal":
                cell.fill = SUBTOTAL_FILL
                cell.font = BOLD
            elif row["_row_type"] == "grandtotal":
                cell.fill = GRANDTOTAL_FILL
                cell.font = WHITE_BOLD
        r += 1

    widths = [10, 12, 14, 10, 10, 14, 16, 14, 16, 18, 14, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w

    if not rows:
        ws["A3"] = "No entries submitted."


def export_to_excel_bytes(latest_entries) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Truck Utilization"
    title = f"Dedicated Truck Utilization Tracker - {date.today().strftime('%d %b %Y')}"
    _populate_worksheet(ws, latest_entries, title)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_day_wise_to_excel_bytes(date_str: str, entries) -> bytes:
    """Builds a single-sheet workbook for a given date's merged (latest-per-plant,
    per-DV-Type) report -- shifts are not segregated in the result, only the most
    recent submission per plant+DV Type is shown."""
    wb = Workbook()
    ws = wb.active
    ws.title = date_str.replace("-", "")[:31]
    title = f"Truck Utilization - {date_str}"
    _populate_worksheet(ws, entries, title)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
